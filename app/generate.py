from PyInquirer import prompt
from examples import custom_style_2
from openpyxl import Workbook

from app.questions import csv_handle, excel, sql, generate_file
from .data import random_data

import json, csv, openpyxl as o

fieldsname = ["Firstname", "Lastname", "Age", "City", "Street", "Zipcode", "Country"]

def generate_txt():
    name, size = new_file()
    d = random_data(size)
    with open(f"bin/{name}.txt", "w") as f:
        f.writelines(
            f'{i.get("Firstname")} {i.get("Lastname")} {i.get("Age")} {i.get("City")} {i.get("Street")} {i.get("Zipcode")} {i.get("Country")}\n' 
            for i in json.loads(d))

def generate_csv():
    c = prompt(csv_handle, style=custom_style_2)
    delim = c.get("csv_option")
    name, size = new_file()
    d = random_data(size)
    
    delimiter = {"comma": ",", "semicolon": ";", "tab": "\t", "vertical bar": "|"}

    with open(f'bin/{name}.csv', "w", encoding='UTF-8') as f:
        writer = csv.DictWriter(f, delimiter=delimiter[delim], fieldnames=fieldsname)
        writer.writeheader()
        writer.writerows(json.loads(d))

def generate_excel():
    e = prompt(excel, style=custom_style_2)
    f = e.get("excel_sheet_name")
    if len(f) == 0:
        f = 'DataSheet'
    name, size = new_file()
    d = random_data(size)

    wb = Workbook()
    sheet = wb.active
    sheet.title = f
    for i in fieldsname:
        sheet.cell(row=1, column=fieldsname.index(i)+1).value = i

    k = 2
    for i in json.loads(d):
        sheet.cell(row=k, column=1).value = i.get("Firstname")
        sheet.cell(row=k, column=2).value = i.get("Lastname")
        sheet.cell(row=k, column=3).value = i.get("Age")
        sheet.cell(row=k, column=4).value = i.get("City")
        sheet.cell(row=k, column=5).value = i.get("Street")
        sheet.cell(row=k, column=6).value = i.get("Zipcode")
        sheet.cell(row=k, column=7).value = i.get("Country")
        k += 1

    wb.save(filename=f'bin/{name}.xlsx')

def generate_sql():
    s = prompt(sql, style=custom_style_2)
    t = s.get("sql_database")
    u = s.get("sql_table")
    v = s.get("sql_column")
    if len(v.strip(",")) != 7:
        v = "Firstname,Lastname,Age,City,Street,Zipcode,Country"
    name, size = new_file()
    d = random_data(size)

    with open(f"bin/{name}.sql", "w") as f:
        f.write(f"use {t}\n")
        f.writelines(
            f'insert into {u} ({v}) values (\"{i.get("Firstname")}\",\"{i.get("Lastname")}\",{i.get("Age")},\"{i.get("City")}\",\"{i.get("Street")}\",\"{i.get("Zipcode")}\",\"{i.get("Country")}\")\n' 
            for i in json.loads(d))

def new_file() -> tuple:
    s = prompt(generate_file, style=custom_style_2)
    a = s.get("file_name")
    b = s.get("size")
    return (a, b)